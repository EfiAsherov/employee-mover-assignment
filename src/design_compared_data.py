import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

def set_width(sheet):
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    for col, value in dims.items():
        sheet.column_dimensions[get_column_letter(col)].width = value
        
def set_style(row, from_column=0, alignment='right', color_font='00000000', bold=False):
    for cell in row[from_column:]:
        cell.alignment = Alignment(horizontal=alignment)
        cell.font = Font(bold=bold, color=color_font)

def seperate_lines(sheet):
    last_row = sheet.max_row
    second_to_last_row = last_row - 1

    if sheet.cell(row=last_row, column=1).value != sheet.cell(row=second_to_last_row, column=1).value:
        for cell in sheet[second_to_last_row]:
            cell.border = Border(bottom=Side(style='thick'))

def initiate_columns_titles(sheet, cells_titles):
    sheet.append(cells_titles)
    set_style(sheet[1], alignment='center', bold=True)

