import openpyxl
from src.employee import Employee
from concurrent.futures import ThreadPoolExecutor


class EmployeeDataExtractor:
    def __init__(self, employees_data_file: str) -> None:
        self.employees_data_file = employees_data_file
        self.employees = {}
    
    def extract_employees_data(self,read_only=False):
        with ThreadPoolExecutor(max_workers=5) as exec:
            file_workbook = openpyxl.load_workbook(self.employees_data_file,read_only=read_only)
            file_sheet = file_workbook.active
            results = [exec.submit(self.create_employee_from_row_data, row) for row in file_sheet.iter_rows(2, file_sheet.max_row)]
        for res in results:
            test_id, employee = res.result()
            self.employees[test_id] = employee
        file_workbook.close
    
    def create_employee_from_row_data(self,row):
        test_id_cell = row[0]
        data_cells = row[1:]
        result = [test_id_cell.value, self.load_employee_data(data_cells)]
        return result
    
    def load_employee_data(self, data_cells):
        return Employee(*[cell.value for cell in data_cells])
